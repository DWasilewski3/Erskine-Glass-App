"""Generate PDF, Excel, and CSV quote files."""

from __future__ import annotations

import base64
import csv
import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa

ROOT = Path(__file__).resolve().parent
LOGO_PNG = ROOT / "static" / "img" / "logo.png"


def _logo_data_uri() -> str:
    if not LOGO_PNG.exists():
        return ""
    encoded = base64.b64encode(LOGO_PNG.read_bytes()).decode("ascii")
    return f"data:image/png;base64,{encoded}"

LINE_COLUMNS = [
    ("qty", "Qty"),
    ("width", "Width"),
    ("height", "Height"),
    ("thick", "Thick"),
    ("type", "Type"),
    ("grid", "Grid"),
    ("color", "Color"),
    ("vert", "VERT"),
    ("hori", "HORI"),
    ("sqft", "SqFt"),
    ("total", "Amount"),
]

NEEDED_COLUMNS = [
    ("qty", "Qty"),
    ("width", "Width"),
    ("height", "Height"),
    ("thick", "Thick"),
    ("type", "Type"),
    ("grid", "Grid"),
    ("color", "Color"),
    ("vert", "VERT"),
    ("hori", "HORI"),
    ("sqft", "SqFt"),
]


def _money(value) -> str:
    try:
        return f"${float(value):,.2f}"
    except (TypeError, ValueError):
        return ""


def _num(value) -> str:
    if value is None or value == "":
        return ""
    try:
        n = float(value)
        return str(int(n)) if n.is_integer() else f"{n:g}"
    except (TypeError, ValueError):
        return str(value)


def render_pdf_html(quote: dict, catalog: dict, *, needed: bool = False) -> str:
    company = catalog.get("company") or {}
    client = quote.get("client") or {}
    logo = _logo_data_uri()
    rows = []
    for line in quote.get("lines") or []:
        amount = "" if needed else f"<td class='r'>{_money(line.get('total'))}</td>"
        rows.append(
            "<tr>"
            f"<td class='c'>{_num(line.get('qty'))}</td>"
            f"<td class='c'>{_num(line.get('width'))} × {_num(line.get('height'))}</td>"
            f"<td class='c'>{line.get('thick') or ''}</td>"
            f"<td>{line.get('type') or ''}</td>"
            f"<td>{line.get('grid') or ''}</td>"
            f"<td>{line.get('color') or ''}</td>"
            f"<td class='c'>{_num(line.get('vert'))}/{_num(line.get('hori'))}</td>"
            f"<td class='c'>{_num(line.get('sqft'))}</td>"
            f"{amount}"
            "</tr>"
        )
    notes = quote.get("notes") or ""
    notes_html = f"<p class='notes'><b>Notes:</b> {notes}</p>" if notes else ""
    logo_html = f"<img class='logo' src='{logo}' />" if logo else ""
    title = "GLASS NEEDED" if needed else "QUOTE"
    amount_th = "" if needed else "<th>Amount</th>"
    empty_cols = "8" if needed else "9"
    if needed:
        totals = (
            f"Qty {_num(quote.get('qty_total'))}"
            f"&nbsp;&nbsp; SqFt {_num(quote.get('sqft_total'))}"
        )
        foot = (
            "Please use this list for the glass order. "
            f"Call {company.get('phone') or ''} with questions."
        )
    else:
        totals = (
            f"Qty {_num(quote.get('qty_total'))}"
            f"&nbsp;&nbsp; SqFt {_num(quote.get('sqft_total'))}<br/>"
            f"<span class='grand'>Total <span class='gold'>{_money(quote.get('grand_total'))}</span></span>"
        )
        foot = (
            "Thank you for the opportunity to quote your glass. This quote is valid for 30 days. "
            f"Call {company.get('phone') or ''} for questions."
        )
    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8" />
<style>
  @page {{ size: letter; margin: 0.55in; }}
  body {{ font-family: Helvetica, Arial, sans-serif; color: #000000; font-size: 10pt; }}
  .header {{ background-color: #1A1612; color: #F4F1EA; padding: 14px 16px; }}
  .logo {{ height: 58px; }}
  .co {{ font-size: 16pt; color: #C4A35A; font-weight: bold; }}
  .tag {{ font-size: 9pt; color: #000000; margin-top: 8px; }}
  .contact {{ font-size: 8.5pt; color: #000000; margin-top: 2px; }}
  h1 {{ font-size: 16pt; color: #000000; margin: 16px 0 8px 0; letter-spacing: 1px; }}
  .meta td {{ padding: 2px 18px 2px 0; font-size: 10pt; color: #000000; }}
  table.lines {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
  table.lines th {{
    background-color: #1A1612; color: #F4F1EA; font-size: 8pt;
    padding: 6px 5px; text-align: left;
  }}
  table.lines td {{ border-bottom: 1px solid #B8B0A4; padding: 6px 5px; font-size: 9pt; color: #000000; }}
  .c {{ text-align: center; }}
  .r {{ text-align: right; }}
  .totals {{ margin-top: 12px; width: 100%; }}
  .totals td {{ padding: 4px 0; color: #000000; }}
  .grand {{ font-size: 13pt; font-weight: bold; color: #000000; }}
  .gold {{ color: #C4A35A; }}
  .notes {{ margin-top: 16px; font-size: 9pt; color: #000000; }}
  .foot {{ margin-top: 22px; font-size: 8pt; color: #000000; border-top: 2px solid #C4A35A; padding-top: 8px; }}
</style>
</head>
<body>
  <div class="header">
    <table width="100%"><tr>
      <td width="90">{logo_html}</td>
      <td>
        <div class="co">{company.get("name") or "Erskine &amp; Sons"}</div>
      </td>
    </tr></table>
  </div>
  <div class="tag">{company.get("tagline") or ""}</div>
  <div class="contact">
    {company.get("phone") or ""}
    · {company.get("email") or ""}
    · {company.get("city") or ""}
    · {company.get("website") or ""}
  </div>
  <h1>{title}</h1>
  <table class="meta">
    <tr>
      <td><b>Customer</b><br/>{client.get("name") or ""}</td>
      <td><b>Date</b><br/>{quote.get("date") or ""}</td>
      <td><b>Quote #</b><br/>{quote.get("quote_number") or ""}</td>
    </tr>
    <tr>
      <td><b>Phone</b><br/>{client.get("phone") or "-"}</td>
      <td colspan="2"><b>Address</b><br/>{client.get("address") or "-"}</td>
    </tr>
  </table>
  <table class="lines">
    <thead>
      <tr>
        <th>Qty</th><th>W × H</th><th>Thick</th><th>Type</th><th>Grid</th>
        <th>Color</th><th>V/H</th><th>SqFt</th>{amount_th}
      </tr>
    </thead>
    <tbody>
      {"".join(rows) or f"<tr><td colspan='{empty_cols}'>No line items</td></tr>"}
    </tbody>
  </table>
  <table class="totals">
    <tr>
      <td></td>
      <td class="r" width="240">
        {totals}
      </td>
    </tr>
  </table>
  {notes_html}
  <div class="foot">
    {foot}
  </div>
</body>
</html>
"""


def _html_to_pdf(html: str) -> bytes:
    buf = io.BytesIO()
    result = pisa.CreatePDF(html, dest=buf, encoding="utf-8")
    if result.err:
        raise RuntimeError("PDF generation failed.")
    return buf.getvalue()


def build_pdf(quote: dict, catalog: dict) -> bytes:
    return _html_to_pdf(render_pdf_html(quote, catalog, needed=False))


def build_glass_needed_pdf(quote: dict, catalog: dict) -> bytes:
    return _html_to_pdf(render_pdf_html(quote, catalog, needed=True))


def build_csv(quote: dict) -> str:
    buf = io.StringIO()
    client = quote.get("client") or {}
    writer = csv.writer(buf)
    writer.writerow(["Erskine & Sons Quote"])
    writer.writerow(["Customer", client.get("name") or ""])
    writer.writerow(["Phone", client.get("phone") or ""])
    writer.writerow(["Address", client.get("address") or ""])
    writer.writerow(["Date", quote.get("date") or ""])
    writer.writerow(["Quote #", quote.get("quote_number") or ""])
    writer.writerow(["Notes", quote.get("notes") or ""])
    writer.writerow([])
    writer.writerow([label for _, label in LINE_COLUMNS])
    for line in quote.get("lines") or []:
        writer.writerow([line.get(key, "") for key, _ in LINE_COLUMNS])
    writer.writerow([])
    writer.writerow(["Qty total", quote.get("qty_total")])
    writer.writerow(["SqFt total", quote.get("sqft_total")])
    writer.writerow(["Grand total", quote.get("grand_total")])
    return buf.getvalue()


def build_glass_needed_csv(quote: dict) -> str:
    buf = io.StringIO()
    client = quote.get("client") or {}
    writer = csv.writer(buf)
    writer.writerow(["Date", "Name"] + [label for _, label in NEEDED_COLUMNS])
    for line in quote.get("lines") or []:
        writer.writerow(
            [quote.get("date") or "", client.get("name") or ""]
            + [line.get(key, "") for key, _ in NEEDED_COLUMNS]
        )
    return buf.getvalue()


def build_xlsx(quote: dict, catalog: dict) -> bytes:
    company = catalog.get("company") or {}
    client = quote.get("client") or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"

    dark = "1A1612"
    gold = "C4A35A"
    cream = "F4F1EA"
    thin = Border(
        left=Side(style="thin", color="D9D2C5"),
        right=Side(style="thin", color="D9D2C5"),
        top=Side(style="thin", color="D9D2C5"),
        bottom=Side(style="thin", color="D9D2C5"),
    )
    header_fill = PatternFill("solid", fgColor=dark)
    header_font = Font(color="F4F1EA", bold=True, name="Calibri", size=10)
    gold_font = Font(color=gold, bold=True, name="Calibri", size=16)
    title_font = Font(color=cream, name="Calibri", size=11)
    money_format = '"$"#,##0.00'

    ws.merge_cells("A1:K2")
    ws["A1"] = f"{company.get('name') or 'Erskine & Sons'}  —  {company.get('tagline') or ''}"
    ws["A1"].font = gold_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(vertical="center", indent=1)
    for col in range(1, 12):
        ws.cell(1, col).fill = header_fill
        ws.cell(2, col).fill = header_fill
    ws.merge_cells("A3:K3")
    ws["A3"] = (
        f"{company.get('phone') or ''}  ·  {company.get('email') or ''}  ·  "
        f"{company.get('city') or ''}  ·  {company.get('website') or ''}"
    )
    ws["A3"].font = title_font
    ws["A3"].fill = header_fill
    for col in range(1, 12):
        ws.cell(3, col).fill = header_fill

    ws["A5"] = "Customer"
    ws["B5"] = client.get("name") or ""
    ws["D5"] = "Date"
    ws["E5"] = quote.get("date") or ""
    ws["G5"] = "Quote #"
    ws["H5"] = quote.get("quote_number") or ""
    ws["A6"] = "Phone"
    ws["B6"] = client.get("phone") or ""
    ws["D6"] = "Address"
    ws["E6"] = client.get("address") or ""
    ws["A7"] = "Notes"
    ws.merge_cells("B7:K7")
    ws["B7"] = quote.get("notes") or ""

    headers = [label for _, label in LINE_COLUMNS]
    for i, label in enumerate(headers, start=1):
        cell = ws.cell(9, i, label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    row = 10
    for line in quote.get("lines") or []:
        values = [line.get(key, "") for key, _ in LINE_COLUMNS]
        for i, value in enumerate(values, start=1):
            cell = ws.cell(row, i, value if value not in (None, "") else None)
            cell.border = thin
            if i == 11:
                cell.number_format = money_format
        row += 1

    row += 1
    ws.cell(row, 10, "Qty")
    ws.cell(row, 11, quote.get("qty_total") or 0)
    row += 1
    ws.cell(row, 10, "SqFt")
    ws.cell(row, 11, quote.get("sqft_total") or 0)
    row += 1
    ws.cell(row, 10, "Total")
    total_cell = ws.cell(row, 11, quote.get("grand_total") or 0)
    total_cell.font = Font(bold=True, size=14, color=dark)
    total_cell.number_format = money_format

    widths = [8, 10, 10, 10, 22, 12, 12, 8, 8, 10, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_title_rows = "1:9"

    needed = wb.create_sheet("Glass Needed")
    needed_headers = ["Date", "Name"] + [label for _, label in NEEDED_COLUMNS]
    for i, label in enumerate(needed_headers, start=1):
        cell = needed.cell(1, i, label)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    r = 2
    for line in quote.get("lines") or []:
        vals = [quote.get("date") or "", client.get("name") or ""] + [
            line.get(key, "") for key, _ in NEEDED_COLUMNS
        ]
        for i, value in enumerate(vals, start=1):
            cell = needed.cell(r, i, value if value not in (None, "") else None)
            cell.border = thin
        r += 1
    for i in range(1, len(needed_headers) + 1):
        needed.column_dimensions[get_column_letter(i)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
